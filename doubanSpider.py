# -*- coding: UTF-8 -*-
import random
import sys
import time
import urllib.parse
import urllib.request
import urllib.error
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Some User Agents
hds = [
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36',
        'Host': 'book.douban.com',
    },
    {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1',
        'Host': 'book.douban.com',
    },
    {
        'User-Agent': 'Mozilla/5.0 (Linux; Android 8.0; Pixel 2 Build/OPD3.170816.012) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Mobile Safari/537.36',
        'Host': 'book.douban.com',
    }
]


def book_spider(book_tag):
    page_num = 0
    book_list = []
    try_times = 0
    # Crawl all books under the label
    # while (1):
    for page_num in range(3):
        # https://book.douban.com/tag/%E4%B8%AA%E4%BA%BA%E7%AE%A1%E7%90%86?start=0 # For Test
        url = 'https://book.douban.com/tag/' + urllib.parse.quote(book_tag) + '?start=' + str(page_num * 20)
        print(url)
        time.sleep(np.random.rand() * 5)
        # Last Version
        try:
            headers = random.choice(hds)
            req = urllib.request.Request(url, headers=headers)  # Construct a Request object
            source_code = urllib.request.urlopen(req).read().decode()
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e)
            continue
        soup = BeautifulSoup(source_code, 'lxml')
        list_soup = soup.find('ul', {'class': 'subject-list'})
        try_times += 1
        if list_soup is None and try_times > 200:
            break
        for book_info in list_soup.find_all(class_="subject-item"):
            # print(book_info)
            # print("**" * 50)
            title = '《' + book_info.find('h2').find('a').attrs['title'] + '》'
            book_url = book_info.find('h2').find('a').attrs['href']
            desc = book_info.find(class_='pub').string.strip()
            desc_list = desc.split('/')

            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info = '作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:-1])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).string.strip()
            except:
                rating = '0.0'
            try:
                people_num = book_info.find('span', class_='pl').string.strip()[1:-1].strip('人评价')
            except:
                people_num = '0'
            # print(title)
            # # print(desc_list)
            # print(book_url)
            # # print(desc_list[0:-3])
            # print(author_info)
            # print(pub_info)
            # print(rating)
            # print(people_num)
            book_list.append([title, rating, people_num, author_info, pub_info, book_url])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print('Downloading Information From Page %d' % page_num)
    return book_list


def do_spider(book_tag_lists):
    """

    :param book_tag_lists:
    :return: Book information list sorted by label under each label
    """
    book_lists = []
    for book_tag in book_tag_lists:
        book_list = book_spider(book_tag)
        book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists, book_tag_lists):
    """
    Add data to the table
    :param book_lists: book information
    :param book_tag_lists: book tags
    :return: xlsx
    """
    wb = Workbook()
    ws = []
    # Add form based on label
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))
    del wb['Sheet']
    # Add data to the form
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者/译者', '出版社', '链接'])
        url_pos = len(['序号', '书名', '评分', '评价人数', '作者/译者', '出版社', '链接'])
        count = 1
        for bl in book_lists[i]:
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4], bl[5]])
            count += 1
            ws[i].cell(row=count, column=url_pos).hyperlink = bl[5]
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i])
    save_path += '.xlsx'
    # r'C:\Users\Administrator\Desktop\{}' the path of your excel
    wb.save(r'C:\Users\Administrator\Desktop\{}'.format(save_path))


if __name__ == '__main__':
    # book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    # book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    # book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    # book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    # book_tag_lists = ['数学']
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    # book_tag_lists = ['商业','理财','管理']
    # book_tag_lists = ['名著']
    # book_tag_lists = ['科普','经典','生活','心灵','文学']
    # book_tag_lists = ['科幻','思维','金融']
    # book_tag_lists = ['个人管理', '时间管理', '投资', '文化']
    book_tag_lists = ['个人管理', '时间管理']
    book_lists = do_spider(book_tag_lists)
    print_book_lists_excel(book_lists, book_tag_lists)
    print(book_lists)
