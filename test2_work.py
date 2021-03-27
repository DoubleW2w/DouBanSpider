# -*- coding: utf-8 -*-
# @Time     : 0:52
# @Author   : DoubleL2l
# @File     : test2_work.py
# @Software : PyCharm
from openpyxl import Workbook
from openpyxl import worksheet
wb = Workbook()
ws = []
book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
book_lists = [
    ['dfg',2.0,'32','dfg','asdsa','https://book.douban.com/tag/%E4%B8%AA%E4%BA%BA%E7%AE%A1%E7%90%86?start=0&type=T'],
    ['2',3.1,'45','cv','asdg','https://book.douban.com/'],
    ['6',7.2,'81','asd','asda','https://blog.csdn.net/qq_41907769?spm=1011.2124.3001.5343&type=blog']
]

for i in range(len(book_tag_lists)):
    ws.append(wb.create_sheet(title=book_tag_lists[i]))
for i in range(len(book_tag_lists)):
    # 添加表头
    ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社', '链接'])
    count = 1
    for bl in book_lists:
        ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4], bl[5]])
        count += 1
        ws[i].cell(row=count, column=7).hyperlink = bl[5]
save_path = 'book_list'
for i in range(len(book_tag_lists)):
    save_path += ('-' + book_tag_lists[i])
save_path += '.xlsx'
del wb['Sheet']
print(wb.sheetnames)
wb.save(r'D:\{}'.format(save_path))